from __future__ import annotations

import html as html_lib
import io
import json
from datetime import datetime
from pathlib import Path

import streamlit as st
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import PatternFill


from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

USERS = ["ananya", "bodi", "meiling"]
ADMIN_USER = "ananya"


FILES: list[str] = [
    "001", "002", "003", "005", "006", "007",
    "009", "010", "011", "012", "013", "089",
    "090", "091", "092", "093", "094", "095",
    "096", "097", "098", "100", "101", "102",
    "104", "105", "120", "121", "122", "123",
    "125", "129", "130", "131", "152", "156",
]

CODING_SCHEMA: dict = {
    "**Expressions:**": [
        {"id": "expr_personal_belief", "label": "Personal belief", "type": "checkbox"},
        {"id": "expr_personal_mindset", "label": "Personal mindset", "type": "checkbox"},
        {"id": "expr_personal_value", "label": "Personal value", "type": "checkbox"},
    ],
    "**Domains:**": [
        {"id": "domain_self", "label": "Self", "type": "checkbox"},
        {"id": "domain_community", "label": "Community", "type": "checkbox"},
        {"id": "domain_society", "label": "Society", "type": "checkbox"},
        {"id": "domain_transcendent", "label": "Transcendent", "type": "checkbox"},
    ],
    "**Emotional tenor:**": [
        {"id": "emotional_tenor", "label": "Emotional tenor", "type": "select",
         "options": ["— select —", "-2: negative", "-1", "0", "1", "2: positive"]},
    ],
    "**Core themes:**": [
        {"id": "self_determination", "label": "Self-determination", "type": "select",
         "options": ["— select —", "0: low", "1", "2", "3", "4: high"]},
        {"id": "connectedness", "label": "Connectedness", "type": "select",
         "options": ["— select —", "0: low", "1", "2", "3", "4: high"]},
    ],
    "**Additional themes:**": [
        {"id": "theme_self_actualization", "label": "Self-actualization", "type": "checkbox"},
        {"id": "theme_world_awareness", "label": "World-awareness", "type": "checkbox"},
        {"id": "theme_trust", "label": "Trust", "type": "checkbox"},
        {"id": "theme_acceptance", "label": "Acceptance", "type": "checkbox"},
        {"id": "theme_intergenerativity", "label": "Intergenerativity", "type": "checkbox"},
    ],
}

USER_SCHEMA_IDS: dict[str, list[str]] = {
    "bodi": [
        "expr_personal_belief", "expr_personal_mindset", "expr_personal_value",
        "domain_self", "domain_community", "domain_society", "domain_transcendent",
        "emotional_tenor"
    ],
    "meiling": [
        "self_determination", "connectedness", "theme_self_actualization",
        "theme_world_awareness", "theme_trust", "theme_acceptance",
        "theme_intergenerativity"
    ],
}

INDIVIDUAL_SECTIONS = {"LIFE CHAPTERS"}

creds = Credentials.from_service_account_info(
        dict(st.secrets["google_service_account"]),
        scopes=["https://www.googleapis.com/auth/drive"],
    )
svc = build("drive", "v3", credentials=creds)


# ── Schema helpers ─────────────────────────────────────────────────────────────

def schema_for_user(annotator: str) -> dict:
    ids = USER_SCHEMA_IDS.get(annotator)
    if annotator == ADMIN_USER or not ids:
        return CODING_SCHEMA
    allowed = set(ids)
    return {
        cat: [item for item in items if item["id"] in allowed]
        for cat, items in CODING_SCHEMA.items()
        if any(item["id"] in allowed for item in items)
    }


def flat_schema(schema: dict | None = None) -> list[dict]:
    s = schema if schema is not None else CODING_SCHEMA
    return [item for items in s.values() for item in items]


def annotator_for_code(code_id: str) -> str | None:
    for user, ids in USER_SCHEMA_IDS.items():
        if code_id in ids:
            return user
    return None


# ── Google Drive helpers ───────────────────────────────────────────────────────

def find_file_in_folder(folder_id, filename):
    query = f"name='{filename}' and '{folder_id}' in parents and trashed=false"
    results = svc.files().list(q=query, fields="files(id, name)").execute()
    files = results.get("files", [])
    return files[0]['id'] if files else None


def _docx_index(filename) -> dict[str, str]:
    """Map docx stem → Drive file_id, cached for the session."""
    if f"docx_{filename}" not in st.session_state:
        file_id = find_file_in_folder(st.secrets["gdrive"]["docs_folder_id"], filename)
        st.session_state[f"docx_{filename}"] = file_id
    return st.session_state[f"docx_{filename}"]

def download_bytes(file_id: str) -> bytes:
    req = svc.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()

def _get_user_data(annotator: str) -> dict:
    """Download and cache the single JSON for this user (all their files)."""
    cache_key = f"user_data_{annotator}"
    if cache_key not in st.session_state:
        jname = f"{annotator}.json"
        file_id = find_file_in_folder(st.secrets["gdrive"]["ann_folder_id"], jname)
        st.session_state[f"{annotator}_file_id"] = file_id
        st.session_state[cache_key] = json.loads(download_bytes(file_id))
    return st.session_state[cache_key]

def _get_ann_data(annotator: str, filename: str) -> dict:
    """Return the per-file section from the user's JSON."""
    return _get_user_data(annotator).get("files", {}).get(filename, {})


def preload_annotations(users: list[str]) -> None:
    """Download each user's single JSON (one request per user)."""
    for user in users:
        _get_user_data(user)


def get_docx_bytes(filename: str) -> bytes:
    idx = _docx_index(f"{filename}.docx")
    if idx not in st.session_state:
        st.session_state[idx] = download_bytes(idx)
    return st.session_state[idx]

def _ann_file_index() -> dict[str, str]:
    """Map filename (e.g. 'ananya.json') → Drive file_id in the annotations folder."""
    if "ann_file_index" not in st.session_state:
        files = gdrive.list_files(st.secrets["gdrive"]["ann_folder_id"])
        st.session_state.ann_file_index = {f["name"]: f["id"] for f in files}
    return st.session_state.ann_file_index


def flush_to_drive(
    annotator: str, filename: str, annotations: dict,
    highlights: list[dict], current_idx: int | None = None,
) -> None:
    """Write the current user JSON to Drive, then update combined.xlsx if needed."""
    save_annotations(annotator, filename, annotations, highlights)

    user_data = _get_user_data(annotator)
    json_bytes = json.dumps(user_data, indent=2).encode("utf-8")
    media = MediaIoBaseUpload(io.BytesIO(json_bytes), mimetype="application/json", resumable=False)
    svc.files().update(fileId=st.session_state[f"{annotator}_file_id"], media_body=media).execute()

    if is_annotated(annotations.get(str(current_idx), {}), schema_for_user(annotator)):
        generate_combined_excel()


# ── Document parsing ───────────────────────────────────────────────────────────

def is_red_run(run) -> bool:
    try:
        rgb = run.font.color.rgb
        if rgb is None:
            return False
        hex_str = str(rgb)
        r, g, b = int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16)
        return r > 150 and g < 120 and b < 120
    except Exception:
        return False


def build_highlight_html(runs: list[dict]) -> str:
    inner = "".join(
        f"<strong>{html_lib.escape(r['text'])}</strong>" if r["bold"]
        else html_lib.escape(r["text"])
        for r in runs
    )
    return (
        f'<mark style="font-size: 18px;background:#fcfb90;padding:2px 5px;border-radius:3px">'
        f"{inner}</mark>"
    )


def extract_highlights(filename: str) -> list[dict]:
    doc = Document(io.BytesIO(get_docx_bytes(filename)))
    highlights = []
    current_title = ""

    for para in doc.paragraphs:
        has_highlight = any(run.font.highlight_color is not None for run in para.runs)
        red_text = "".join(run.text for run in para.runs if is_red_run(run)).strip()
        if red_text and not has_highlight:
            current_title = red_text
        if not has_highlight:
            continue

        current_runs: list[dict] = []
        groups: list[list[dict]] = []
        for run in para.runs:
            if run.text == ' ':
                current_runs.append({"text": run.text, "bold": False})
            elif run.font.highlight_color is not None:
                current_runs.append({"text": run.text, "bold": bool(run.bold)})
            else:
                if current_runs:
                    groups.append(current_runs)
                    current_runs = []
        if current_runs:
            groups.append(current_runs)

        for runs in groups:
            plain = "".join(r["text"] for r in runs)
            if plain.strip():
                highlights.append({"runs": runs, "text": plain,
                                   "title": current_title, "context": para.text})
    return group_highlights(highlights)


def group_highlights(flat: list[dict]) -> list[dict]:
    units: list[dict] = []
    i = 0
    while i < len(flat):
        h = flat[i]
        title = h["title"]
        if any(s in title.upper() for s in INDIVIDUAL_SECTIONS):
            units.append({"id": len(units), "title": title, "items": [h], "text": h["text"]})
            i += 1
        else:
            group = [h]
            j = i + 1
            while j < len(flat) and flat[j]["title"] == title:
                group.append(flat[j])
                j += 1
            units.append({
                "id": len(units), "title": title,
                "items": group, "text": " ".join(g["text"] for g in group),
            })
            i = j
    return units


# ── Annotation persistence ─────────────────────────────────────────────────────

def load_annotations(annotator: str, filename: str) -> dict:
    return _get_ann_data(annotator, filename).get("annotations", {})


def save_annotations(
    annotator: str, filename: str, annotations: dict, highlights: list[dict],
) -> None:
    """Update in-memory state only. Call flush_to_drive() on navigation to persist."""
    highlights_meta = {
        str(i): {
            "title": h.get("title", ""),
            "quote": " ".join(h["items"][0]["text"].split()[:10])
                     + ("…" if len(h["items"][0]["text"].split()) > 10 else ""),
        }
        for i, h in enumerate(highlights)
    }
    user_data = _get_user_data(annotator)
    user_data.setdefault("files", {})[filename] = {
        "total_highlights": len(highlights),
        "highlights_meta": highlights_meta,
        "annotations": annotations,
    }
    user_data["last_updated"] = datetime.now().isoformat()
    st.session_state[f"user_data_{annotator}"] = user_data


def is_complete(annotator: str, filename: str) -> bool:
    data = _get_ann_data(annotator, filename)
    total = data.get("total_highlights", -1)
    if total < 0:
        return False
    anns = data.get("annotations", {})
    return all(is_annotated(anns.get(str(i), {}), schema_for_user(annotator)) for i in range(total))


def is_annotated(ann: dict, schema: dict | None = None) -> bool:
    if not ann:
        return False
    codes = ann.get("codes", {})
    select_ids = [item["id"] for item in flat_schema(schema) if item["type"] == "select"]
    if select_ids:
        return all(codes.get(sid, "") not in ("", "— select —") for sid in select_ids)
    return bool(codes)


# ── Excel export ───────────────────────────────────────────────────────────────

def _fmt(val) -> str:
    if val is None or val in ("", "— select —"):
        return ""
    if isinstance(val, bool):
        return "1" if val else "0"
    elif ":" in str(val):
        return int(str(val).split(":")[0])
    try:
        return int(val)
    except (ValueError, TypeError):
        return str(val)


def generate_combined_excel() -> None:
    grey_fill   = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Annotations"

    schema_items = flat_schema()
    header: list[str] = ["Participant", "Section", "Quote"]
    col_meta: list[tuple] = []
    for item in schema_items:
        other_user = annotator_for_code(item["id"])
        white_col = len(header) + 1
        temp_label = item["label"][len("Personal "):].title() if "Personal" in item["label"] else item["label"]
        header.append(temp_label)
        grey_col = len(header) + 1
        header.append(f"{temp_label} ({other_user})")
        col_meta.append((item, other_user, white_col, grey_col))
    header.append("Notes")

    ws.append(header)
    for _, _, _, grey_col in col_meta:
        ws.cell(1, grey_col).fill = grey_fill

    for filename in FILES:
        total = 0
        file_data: dict[str, dict] = {}
        for user in USERS:
            d = _get_user_data(user).get("files", {}).get(filename, {})
            total = max(total, d.get("total_highlights", 0))
            file_data[user] = d
        if total == 0:
            continue

        highlights_meta = next(
            (d.get("highlights_meta", {}) for d in file_data.values() if d.get("highlights_meta")),
            {},
        )
        user_anns = {user: file_data[user].get("annotations", {}) for user in USERS}

        for i in range(total):
            meta = highlights_meta.get(str(i), {})
            title = meta.get("title", "")[8:].lower()
            quote = meta.get("quote", "")

            u1_ann = user_anns.get(ADMIN_USER, {}).get(str(i), {})
            u1_complete = is_annotated(u1_ann, schema_for_user(ADMIN_USER))

            notes_parts = [
                f"{user}: {user_anns[user][str(i)]['notes']}"
                for user in USERS
                if user_anns.get(user, {}).get(str(i), {}).get("notes", "").strip()
            ]

            row: list = [filename, title, quote]
            for item, other_user, _, _ in col_meta:
                sid = item["id"]
                u1_val = _fmt(u1_ann.get("codes", {}).get(sid)) if u1_complete else ""
                ou_ann = user_anns.get(other_user, {}).get(str(i), {}) if other_user else {}
                ou_complete = is_annotated(ou_ann, schema_for_user(other_user)) if other_user else False
                ou_val = _fmt(ou_ann.get("codes", {}).get(sid)) if ou_complete else ""
                row += [u1_val, ou_val]
            row.append("\n".join(notes_parts))

            ws.append(row)
            row_idx = ws.max_row

            for item, other_user, white_col, grey_col in col_meta:
                sid = item["id"]
                u1_val = _fmt(u1_ann.get("codes", {}).get(sid)) if u1_complete else ""
                ou_ann = user_anns.get(other_user, {}).get(str(i), {}) if other_user else {}
                ou_complete = is_annotated(ou_ann, schema_for_user(other_user)) if other_user else False
                ou_val = _fmt(ou_ann.get("codes", {}).get(sid)) if ou_complete else ""

                ws.cell(row_idx, grey_col).fill = grey_fill
                if u1_val != "" and ou_val != "" and u1_val != ou_val:
                    ws.cell(row_idx, white_col).fill = yellow_fill
                    ws.cell(row_idx, grey_col).fill = yellow_fill

    buf = io.BytesIO()
    wb.save(buf)
    if "combined_file_id" not in st.session_state:
        st.session_state["combined_file_id"] = find_file_in_folder(st.secrets["gdrive"]["ann_folder_id"], "combined.xlsx")
    media = MediaIoBaseUpload(io.BytesIO(buf.getvalue()), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resumable=False)
    svc.files().update(fileId=st.session_state["combined_file_id"], media_body=media).execute()


# ── Pages ──────────────────────────────────────────────────────────────────────

def render_name_page() -> None:
    _, col, _ = st.columns([1, 2, 1])
    with col:
        st.title("Life Philosophy Coding Interface")
        st.subheader("Get started")
        name = st.text_input("Your name")
        name_valid = name.strip() in USERS
        if name.strip() and not name_valid:
            st.error(f'"{name.strip()}" is not a recognised annotator. Please enter one of the assigned names.')
        if st.button("Continue", disabled=not name_valid, type="primary"):
            st.session_state.annotator = name.strip()
            st.session_state.page = "files"
            st.rerun()


def render_file_page() -> None:
    annotator = st.session_state.annotator
    _, col, _ = st.columns([1, 2, 1])
    with col:
        st.title("Life Philosophy Coding Interface")
        st.subheader(f"Welcome, {annotator}")

        with st.spinner("Loading progress…"):
            users_to_load = USERS if annotator == ADMIN_USER else [annotator]
            preload_annotations(users_to_load)

        n_total = len(FILES)

        if annotator == ADMIN_USER:
            for u in USERS:
                u_done = sum(1 for s in FILES if is_complete(u, s))
                st.progress(u_done / n_total if n_total else 0)
                st.caption(f"{u}: {u_done} of {n_total} files complete")
        else:
            n_done = sum(1 for s in FILES if is_complete(annotator, s))
            st.progress(n_done / n_total if n_total else 0)
            st.caption(f"{n_done} of {n_total} files complete")

        personal_complete = [s for s in FILES if is_complete(annotator, s)]
        personal_incomplete = [s for s in FILES if not is_complete(annotator, s)]
        display_options = personal_incomplete + [f"✓ {s}" for s in personal_complete]
        stem_options = personal_incomplete + personal_complete

        file_choice_display = st.selectbox("Document", display_options)
        file_choice = stem_options[display_options.index(file_choice_display)]

        col_back, col_start = st.columns([1, 2])
        with col_back:
            if st.button("← Back"):
                st.session_state.page = "name"
                st.rerun()
        with col_start:
            if st.button("Start annotating", type="primary"):
                st.session_state.doc_name = file_choice
                st.session_state.page = "annotate"
                st.session_state.pop("highlights", None)
                st.rerun()


def render_annotation() -> None:
    annotator = st.session_state.annotator
    doc_name = st.session_state.doc_name

    if "highlights" not in st.session_state:
        with st.spinner("Loading document…"):
            st.session_state.highlights = extract_highlights(doc_name)
        st.session_state.annotations = load_annotations(annotator, doc_name)
        st.session_state.current_idx = 0

    highlights = st.session_state.highlights
    annotations = st.session_state.annotations
    n = len(highlights)
    user_schema = schema_for_user(annotator)

    if n == 0:
        st.warning("No highlighted sections found in this document.")
        if st.button("← Back"):
            st.session_state.page = "files"
            st.rerun()
        return

    idx = st.session_state.get("current_idx", 0)

    def _go_to_files():
        st.session_state.page = "files"
        for key in ("highlights", "annotations", "current_idx"):
            st.session_state.pop(key, None)
        st.session_state.pop("ann_file_index", None)

    with st.sidebar:
        st.markdown(f"**Interview: {doc_name}**")
        st.caption(f"Annotator: {annotator}")
        completed = sum(1 for v in annotations.values() if is_annotated(v, user_schema))
        st.progress(completed / n)
        st.caption(f"{completed} of {n} sections coded")
        st.divider()
        for i, h in enumerate(highlights):
            saved = annotations.get(str(i), {})
            has_label = is_annotated(saved, user_schema)
            icon = "✅" if has_label else "○"
            nav_title = h["title"] if h["title"] else h["text"]
            label_text = nav_title.title()[8:]
            current_marker = "▶ " if i == idx else ""
            if st.button(f"{current_marker}{icon} {i + 1}. {label_text}",
                         key=f"nav_{i}", use_container_width=True):
                flush_to_drive(annotator, doc_name, annotations, highlights, current_idx=idx)
                st.session_state.current_idx = i
                st.rerun()
        st.divider()
        if st.button("← Change document"):
            flush_to_drive(annotator, doc_name, annotations, highlights, current_idx=idx)
            _go_to_files()
            st.rerun()

    highlight = highlights[idx]
    saved = annotations.get(str(idx), {})

    para_groups: list[tuple[str, list[dict]]] = []
    ctx_index: dict[str, int] = {}
    for item in highlight["items"]:
        ctx = item["context"]
        if ctx not in ctx_index:
            ctx_index[ctx] = len(para_groups)
            para_groups.append((ctx, [item]))
        else:
            para_groups[ctx_index[ctx]][1].append(item)

    for p_idx, (context, items) in enumerate(para_groups):
        pos = 0
        parts: list[str] = []
        for item in items:
            start = context.find(item["text"], pos)
            if start >= 0:
                parts.append(html_lib.escape(context[pos:start]).replace("\n\n", "</span>\n\n<span style='font-size: 18px;'>"))
                parts.append(build_highlight_html(item["runs"]))
                pos = start + len(item["text"])
            else:
                parts.append(build_highlight_html(item["runs"]))
        parts.append(html_lib.escape(context[pos:]).replace("\n\n", "</span>\n\n<span style='font-size: 18px;'>"))
        with st.container(height=500):
            if highlight["title"]:
                st.markdown(f"#### {idx + 1}. {html_lib.escape(highlight['title'])}", unsafe_allow_html=True)
            st.markdown(f'<span style="font-size: 18px;">{"".join(parts)}</span>', unsafe_allow_html=True)
            if p_idx < len(para_groups) - 1:
                st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("<div style='height: 30px;'></div>", unsafe_allow_html=True)
    saved_codes = saved.get("codes", {})
    new_codes: dict = {}
    cols = st.columns(5)
    for col_id, (category, items) in enumerate(user_schema.items()):
        with cols[col_id]:
            st.markdown(category, unsafe_allow_html=True)
            for schema_item in items:
                #with cols[col_id+1]:
                sid = schema_item["id"]
                if schema_item["type"] == "checkbox":
                    new_codes[sid] = st.checkbox(
                        schema_item["label"],
                        value=bool(saved_codes.get(sid, False)),
                        key=f"code_{idx}_{sid}",
                    )
                else:
                    opts = schema_item["options"]
                    saved_val = saved_codes.get(sid, opts[0])
                    opt_idx = opts.index(saved_val) if saved_val in opts else 0
                    new_codes[sid] = st.selectbox(
                        schema_item["label"], opts, index=opt_idx,
                        key=f"code_{idx}_{sid}",
                    )

    notes = st.text_area("Notes (Optional)", value=saved.get("notes", ""), height=100, key=f"notes_{idx}")

    first_words = highlight["items"][0]["text"].split()
    quote = " ".join(first_words[:10]) + ("…" if len(first_words) > 10 else "")
    annotations[str(idx)] = {
        "codes": new_codes, "notes": notes,
        "text": highlight["text"], "title": highlight["title"], "quote": quote,
    }
    # Update in-memory state on every rerun (no Drive I/O)
    save_annotations(annotator, doc_name, annotations, highlights)

    col_prev, _, col_next = st.columns([1, 4, 1])
    with col_prev:
        if st.button("← Prev", disabled=idx == 0):
            flush_to_drive(annotator, doc_name, annotations, highlights, current_idx=idx)
            st.session_state.current_idx = idx - 1
            st.rerun()
    with col_next:
        if idx < n - 1:
            if st.button("Next →"):
                flush_to_drive(annotator, doc_name, annotations, highlights, current_idx=idx)
                st.session_state.current_idx = idx + 1
                st.rerun()
        else:
            if st.button("Done ✓", type="primary"):
                flush_to_drive(annotator, doc_name, annotations, highlights, current_idx=idx)
                _go_to_files()
                st.rerun()


def main():
    st.set_page_config(page_title="Life Philosophy Coding Interface", layout="wide")
    st.html(
        """
        <style>
        [data-testid="stCheckbox"] [data-testid="stWidgetLabel"] p {
            font-size: 18px !important;
        }
        div[data-testid="stSelectbox"] label p {
            font-size: 18px !important;
        }
        [data-testid="stVerticalBlock"] {
            gap: 0.1rem; 
        }
        </style>
        """
    )

    if "page" not in st.session_state:
        st.session_state.page = "name"
    page = st.session_state.page
    if page == "name":
        render_name_page()
    elif page == "files":
        render_file_page()
    else:
        render_annotation()


if __name__ == "__main__":
    main()
